#!/usr/bin/env python3
"""Generate NCBI GenBank submission Excel template (.xlsx).

Produces scripts/docs/ncbi_submission_template.xlsx with 3 sheets:
  1. BankIt_Source_Modifiers — new submissions (Sequence_ID first column)
  2. Record_Update — update existing records (acc. num. first column)
  3. 字段说明 — field descriptions with format requirements

Usage:
    python scripts/create_ncbi_template.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = Path(__file__).parent / "docs" / "ncbi_submission_template.xlsx"

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, size=11)
DATA_FONT = Font(name="Arial", size=11)
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
REQUIRED_FILL = PatternFill("solid", fgColor="FFFFCC")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

# ── BankIt columns (new submission) ─────────────────────────────────────────
BANKIT_COLS = [
    "Sequence_ID",
    "Collected_by",
    "Collection_date",
    "Country (geo_loc_name)",
    "Isolation_source",
    "Isolate",
    "Lat_Lon",
    "Specimen_voucher",
    "Host",
    "Culture_collection",
    "Clone",
    "Strain",
    "Altitude",
    "Note",
    "Sex",
    "Dev_stage",
    "Haplotype",
    "Genotype",
    "Tissue_type",
    "Cell_line",
    "Fwd_primer_name",
    "Fwd_primer_seq",
    "Rev_primer_name",
    "Rev_primer_seq",
]
BANKIT_REQUIRED = {"Sequence_ID", "Collection_date", "Country (geo_loc_name)"}

BANKIT_SAMPLE = [
    ["Seq1", "C. Grant", "31-Jan-2001", "USA", "soil", "A", "13.57 N 24.68 W", "MKP 334"],
    ["Seq2", "S. Tracy", "28-Feb-2002", "Slovakia", "contaminated soil", "B", "13.24 N 24.35 W", "MKP 1230"],
    ["Seq3", "C. Grant", "31-Jan-2001", "USA", "soil", "", "13.57 N 24.68 W", "MKP 335"],
]

# ── Record Update columns ───────────────────────────────────────────────────
UPDATE_COLS = [
    "acc. num.",
    "geo_loc_name",
    "Lat_Lon",
    "Collection_date",
    "Collected_by",
    "Specimen_voucher",
    "Isolation_source",
]
UPDATE_REQUIRED = {"acc. num."}

UPDATE_SAMPLE = [
    ["AB123456", "Norway: Hordaland, Bergen", "60.39 N 5.33 E", "08-Nov-2005", "Smith, J.", "SMNH:211581", "freshwater sediment"],
    ["AB123457", "China: Sichuan, Chengdu", "30.57 N 104.07 E", "1993", "Wang, L.", "personal:Li Wang:2020-001", "soil"],
]

# ── Field descriptions ──────────────────────────────────────────────────────
FIELD_DESC = [
    # (Field, Required, Format, Example, Notes)
    ("Sequence_ID", "必填", "自由文本", "Seq1", "BankIt 新提交用。每行唯一标识符"),
    ("acc. num.", "必填", "GenBank 登录号", "AB123456", "更新已有记录用。不含版本号"),
    ("Country (geo_loc_name)", "必填", "Country: Region, Locality", "Norway: Hordaland, Bergen", "BankIt 提交列名。冒号后跟省/州，逗号分隔更细地域"),
    ("geo_loc_name", "必填", "Country: Region, Locality", "Norway: Hordaland, Bergen", "记录更新列名。与 Country (geo_loc_name) 格式相同"),
    ("Lat_Lon", "推荐", "dd.dd N|S ddd.dd E|W", "57.68 N 11.96 E", "十进制小数，保留 2-4 位。不接受度分秒"),
    ("Collection_date", "必填", "DD-Mon-YYYY / Mon-YYYY / YYYY", "08-Nov-2005", "月份缩写: Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec"),
    ("Collected_by", "推荐", "Lastname, F.", "Smith, J.; Doe, A.", "多人用分号分隔。不缩写为 et al."),
    ("Specimen_voucher", "推荐", "institution:specimen-id", "SMNH:211581", "格式: institution-code[:collection-code]:specimen-id"),
    ("Isolation_source", "可选", "自由文本", "soil", "样本来源描述"),
    ("Isolate", "可选", "自由文本", "strain ABC123", "分离株名称"),
    ("Host", "可选", "自由文本", "Homo sapiens", "宿主生物"),
    ("Culture_collection", "可选", "institution:accession", "ATCC:12345", "培养物保藏编号"),
    ("Clone", "可选", "自由文本", "clone xyz1", "克隆名称"),
    ("Strain", "可选", "自由文本", "strain K-12", "菌株名称"),
    ("Altitude", "可选", "value unit", "1500 m", "海拔高度"),
    ("Note", "可选", "自由文本", "附加说明", ""),
    ("Sex", "可选", "自由文本", "female", ""),
    ("Dev_stage", "可选", "自由文本", "larva", "发育阶段"),
    ("Haplotype", "可选", "自由文本", "H1", ""),
    ("Genotype", "可选", "自由文本", "GT1", ""),
    ("Tissue_type", "可选", "自由文本", "liver", ""),
    ("Cell_line", "可选", "自由文本", "HeLa", ""),
    ("Fwd_primer_name", "可选", "自由文本", "LCO1490", "正向引物名称"),
    ("Fwd_primer_seq", "可选", "碱基序列", "GGTCAACAAATCATAAAGATATTGG", "正向引物序列"),
    ("Rev_primer_name", "可选", "自由文本", "HCO2198", "反向引物名称"),
    ("Rev_primer_seq", "可选", "碱基序列", "TAAACTTCAGGGTGACCAAAAAATCA", "反向引物序列"),
]

# 填充色按必填程度
REQ_FILLS = {"必填": PatternFill("solid", fgColor="FFFFCC"), "推荐": PatternFill("solid", fgColor="E2EFDA"), "可选": None}


def _write_sheet(ws, columns, required_cols, sample_rows):
    """Write a data sheet with headers and sample rows."""
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = REQUIRED_FILL if col in required_cols else HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = WRAP

    for ri, row_data in enumerate(sample_rows, 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP

    # Auto-width
    for ci, col in enumerate(columns, 1):
        max_len = len(str(col))
        for ri, row_data in enumerate(sample_rows, 2):
            if ci - 1 < len(row_data) and row_data[ci - 1]:
                max_len = max(max_len, len(str(row_data[ci - 1])))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_desc_sheet(ws):
    """Write the field descriptions sheet."""
    headers = ["字段名", "必填程度", "格式", "示例", "说明"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = WRAP

    for ri, (field, req, fmt, example, notes) in enumerate(FIELD_DESC, 2):
        vals = [field, req, fmt, example, notes]
        fill = REQ_FILLS.get(req)
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP
            if fill:
                cell.fill = fill

    widths = [28, 10, 32, 32, 40]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    ws.freeze_panes = "A2"


def main():
    wb = Workbook()

    # Sheet 1: BankIt
    ws1 = wb.active
    ws1.title = "BankIt_Source_Modifiers"
    _write_sheet(ws1, BANKIT_COLS, BANKIT_REQUIRED, BANKIT_SAMPLE)

    # Sheet 2: Record Update
    ws2 = wb.create_sheet("Record_Update")
    _write_sheet(ws2, UPDATE_COLS, UPDATE_REQUIRED, UPDATE_SAMPLE)

    # Sheet 3: Field descriptions
    ws3 = wb.create_sheet("字段说明")
    _write_desc_sheet(ws3)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"NCBI 提交模板已生成: {OUTPUT}")
    print(f"  Sheet 1: BankIt_Source_Modifiers ({len(BANKIT_COLS)} 列, {len(BANKIT_SAMPLE)} 行示例)")
    print(f"  Sheet 2: Record_Update ({len(UPDATE_COLS)} 列, {len(UPDATE_SAMPLE)} 行示例)")
    print(f"  Sheet 3: 字段说明 ({len(FIELD_DESC)} 个字段)")


if __name__ == "__main__":
    main()
